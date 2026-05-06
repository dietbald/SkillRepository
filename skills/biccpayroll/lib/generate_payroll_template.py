#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BICC Payroll Generation — TEMPLATE.

This file is meant to be COPIED and EDITED per period. The "EDIT PER PERIOD"
sections at the top capture everything Tina submits in the Payroll Input
template plus the period dates. The generation logic below is stable and
should not need changes.

USAGE:
1. Copy this file to your work dir, e.g. `generate_payroll_<paydate>.py`
2. Read Tina's Payroll Input xlsx for this period and update the EDIT PER PERIOD
   sections below
3. Run: `python generate_payroll_<paydate>.py`
4. Inspect the verification output. If FAIL, fix and re-run.
"""

import os
import sys
import shutil
import subprocess
import time
import json
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import win32com.client
from openpyxl.utils import get_column_letter

# Import shared config
sys.path.insert(0, str(Path(__file__).parent))
from biccpayroll_config import (
    DAILY_EMPLOYEES, MONTHLY_EMPLOYEES, EXEMPT_EMPLOYEES, HABARADAS_LIKE,
    EXCLUDED_NAMES, RESIGNED_ZERO, ALREADY_ZEROED, RESIGNED_ROW_MAP,
    V_FORMULA_ROWS, PASSWORD, INCENTIVE_AMOUNT, LATE_THRESHOLD_MIN, LATE_COUNT_LIMIT,
    normalize_surname, get_employee_type, surname_from_payroll,
)

# ============================================================================
# ============================================================================
# EDIT PER PERIOD — everything below until the next big banner
# ============================================================================
# ============================================================================

# Period
START = date(2026, 4, 16)
END = date(2026, 4, 30)
HOLIDAYS = {
    # date(2026, 4, 9): "regular",  # example
}

# File paths
WORK_DIR = "C:/Users/thiba/Documents/Payroll/work"
ATTENDANCE_DIR = "C:/Users/thiba/BICC/HR Payroll - Documents/Admin Payroll 2026/20260502_Payroll May 10 Cutoff"
TEMPLATE = os.path.join(WORK_DIR, "current_payroll_apr25_dec.xlsx")  # prior decrypted file
ATTENDANCE_FILE = os.path.join(ATTENDANCE_DIR,
    "BICC_Admin Attendance Sheet - April 16 - 30, 2026.xlsx")
ATTENDANCE_JSON = os.path.join(WORK_DIR, "attendance_may10_results.json")
PRIOR_JSON = os.path.join(WORK_DIR, "attendance_apr25_results.json")  # for incentive eval
OUTPUT_DEC = os.path.join(WORK_DIR, "current_payroll_may10_dec.xlsx")
OUTPUT_ENC = os.path.join(WORK_DIR, "current_payroll_may10.xlsx")
OUTPUT_FINAL = os.path.join(ATTENDANCE_DIR,
    "New_Payroll_2026-04-16_to_2026-04-30_PayDate_2026-05-10.xlsx")
OLD_SHEET = "Payroll April 25"
NEW_SHEET = "Payroll May 10"

# From Payroll Input — Section 1B Resignations (mid-period LWD)
LWD_RESIGNATIONS = {
    # "esmeralda": date(2026, 4, 25),
    # "lapizar":   date(2026, 4, 25),
}

# From Payroll Input — Section 3 Adjustments (W column = col 23)
# CHECK FIRST: read EI for affected employees to ensure no double-count with prior EI changes
ADJUSTMENTS = {
    # "cabunagan":   -550,
    # "lood":         100,
}

# From Payroll Input — Section 4 Cash Advances (AF column = col 32)
CASH_ADVANCES = {
    # "maquiling": 1000,
}

# From Payroll Input — Section 2 Loan Updates
SSS_LOANS = {
    # "obillos": 1809.09,
}
HDMF_LOANS = {
    # "obillos": 1611.97,
}

# From Payroll Input — Section 1D bank changes
BANK_UPDATES = {
    # "prologo": "1479369026",
}

# Holiday premium workers (MONTHLY/DAILY only). Map surname to {regular: count, special: count}
HOLIDAY_PREMIUM = {
    # "futotana": {"regular": 1},
}

# V (allowance) overrides — for mid-period resignations, prorate by calendar days of month
# Formula: monthly_allowance × (calendar_days_employed / total_calendar_days_in_month)
V_OVERRIDES = {
    # "lapizar": 2500,  # 3000 × 25/30
}

# ============================================================================
# ============================================================================
# STABLE LOGIC — no edits below per period
# ============================================================================
# ============================================================================

all_dates = [START + timedelta(days=i) for i in range((END - START).days + 1)]
workdays = [d for d in all_dates if d.weekday() < 6]
PERIOD_DAYS = len(workdays)

# Determine cutoff
CUTOFF = "A" if START.day == 1 else "B"
D4_VALUE = "01-15" if CUTOFF == "A" else "16-31"

print(f"Period: {START} to {END} | Cutoff {CUTOFF} | Period Days: {PERIOD_DAYS}")
print(f"D4 value: {D4_VALUE}")
print(f"Holidays: {dict(HOLIDAYS) or '(none)'}")
print(f"LWD resignations: {LWD_RESIGNATIONS or '(none)'}")
sys.stdout.flush()


def find_json_key(att_data, surname):
    sn_norm = normalize_surname(surname)
    if surname in att_data:
        return surname
    for key in att_data:
        if normalize_surname(key) == sn_norm:
            return key
    return None


# STEP 1: Load attendance JSONs
print(f"\n{'='*80}\nSTEP 1: Load Attendance JSON\n{'='*80}")
with open(ATTENDANCE_JSON) as f:
    attendance_data = json.load(f)
if "0" in attendance_data:
    del attendance_data["0"]

prior_data = {}
if Path(PRIOR_JSON).exists():
    with open(PRIOR_JSON) as f:
        prior_data = json.load(f)
    if "0" in prior_data:
        del prior_data["0"]
print(f"  Loaded {len(attendance_data)} current + {len(prior_data)} prior")
sys.stdout.flush()


# STEP 2: Read template state via Excel COM
print(f"\n{'='*80}\nSTEP 2: Read Template State\n{'='*80}")
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
wb_com = excel.Workbooks.Open(os.path.abspath(TEMPLATE))
time.sleep(2)
excel.CalculateFull()
time.sleep(1)

ws_att_com = wb_com.Sheets("ATTENDANCE")
att_names = {}
att_1st_late = {}
att_1st_awol = {}
att_surname_to_row = {}
for r in range(9, 43):
    name = ws_att_com.Cells(r, 4).Value
    if name and isinstance(name, str) and name.strip():
        att_names[r] = name.strip()
        att_surname_to_row[surname_from_payroll(name)] = r
        att_1st_late[r] = float(ws_att_com.Cells(r, 16).Value or 0)
        att_1st_awol[r] = float(ws_att_com.Cells(r, 19).Value or 0)
print(f"  ATTENDANCE: {len(att_names)} employees")

ws_pay_com = wb_com.Sheets(OLD_SHEET)
payroll_surname_to_row = {}
for r in range(9, 46):
    name = ws_pay_com.Cells(r, 3).Value
    if name and isinstance(name, str) and name.strip():
        payroll_surname_to_row[surname_from_payroll(name)] = r
print(f"  Payroll: {len(payroll_surname_to_row)} employees")

ws_bd_com = wb_com.Sheets("Bank Details")
bd_total_row = None
bd_oresco_row = None
for r in range(1, 60):
    val_b = ws_bd_com.Cells(r, 2).Value
    if val_b and isinstance(val_b, str) and "TOTAL" in val_b.upper():
        bd_total_row = r
    for c in [3, 5, 6]:
        v = ws_bd_com.Cells(r, c).Value
        if v and isinstance(v, str) and "oresco" in v.lower():
            bd_oresco_row = r
            break

wb_com.Close(False)
excel.Quit()
del excel
time.sleep(1)


# STEP 3: Copy template & rename sheet
print(f"\n{'='*80}\nSTEP 3: Copy & Rename\n{'='*80}")
shutil.copy2(TEMPLATE, OUTPUT_DEC)
wb = openpyxl.load_workbook(OUTPUT_DEC)
ws_p = wb[OLD_SHEET]
ws_p.title = NEW_SHEET
print(f"  Renamed: '{OLD_SHEET}' -> '{NEW_SHEET}'")

# Replace cross-sheet refs
old_ref = f"'{OLD_SHEET}'"
new_ref = f"'{NEW_SHEET}'"
replaced = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and old_ref in cell.value:
                cell.value = cell.value.replace(old_ref, new_ref)
                replaced += 1
print(f"  Replaced {replaced} cross-sheet refs")


# STEP 4: Update ATTENDANCE
print(f"\n{'='*80}\nSTEP 4: Update ATTENDANCE\n{'='*80}")
ws_att = wb["ATTENDANCE"]
ws_att["C3"] = START.strftime("%B")
ws_att["C4"] = START.strftime("%B")
ws_att["D4"] = D4_VALUE
ws_att["E4"] = START.year
ws_att["C7"] = "First Half (1-15)" if CUTOFF == "A" else "Second Half (16-31)"

for att_row, att_name in att_names.items():
    sn_norm = normalize_surname(surname_from_payroll(att_name))

    # Excluded / exempt / habaradas: zero attendance, preserve 1st-half if Cutoff B
    is_excluded_role = (
        sn_norm in {normalize_surname(x) for x in EXCLUDED_NAMES}
        or sn_norm in {normalize_surname(x) for x in EXEMPT_EMPLOYEES}
        or sn_norm in {normalize_surname(x) for x in HABARADAS_LIKE}
    )
    if is_excluded_role:
        for c in [7, 8, 9, 10, 11, 12, 13]:
            ws_att.cell(att_row, c, 0)
        if CUTOFF == "B":
            ws_att.cell(att_row, 17, 0)
            ws_att.cell(att_row, 18, att_1st_late[att_row])
            ws_att.cell(att_row, 20, 0)
            ws_att.cell(att_row, 21, att_1st_awol[att_row])
            ws_att.cell(att_row, 22, 0)
        continue

    json_key = find_json_key(attendance_data, sn_norm)
    if not json_key:
        print(f"  WARN: no attendance for {att_name}")
        continue
    att = attendance_data[json_key]

    # CRITICAL: DAILY ATTENDANCE L (col 12) = 0 to prevent S=J*Q double-deduction
    abs_for_payroll = 0 if att.get("is_daily") else att["abs"]

    ws_att.cell(att_row, 7, att["days"])         # G: Days
    ws_att.cell(att_row, 8, att["tardy_hrs"])    # H: Tardy hrs
    ws_att.cell(att_row, 9, att["lwp"])          # I: LWP
    ws_att.cell(att_row, 10, att["lwop"])        # J: LWOP
    ws_att.cell(att_row, 11, att["awol"])         # K: AWOL
    ws_att.cell(att_row, 12, abs_for_payroll)    # L: Total Abs
    ws_att.cell(att_row, 13, att["ot"])           # M: OT hrs (already 0 for FIXED via compute)

    # Incentive cols P-V
    if CUTOFF == "A":
        ws_att.cell(att_row, 16, att["tardy_hrs"])  # P
        ws_att.cell(att_row, 17, 0)                  # Q
        ws_att.cell(att_row, 18, att["tardy_hrs"])  # R
        ws_att.cell(att_row, 19, att["awol"])         # S
        ws_att.cell(att_row, 20, 0)                  # T
        ws_att.cell(att_row, 21, att["awol"])         # U
        ws_att.cell(att_row, 22, 0)                  # V (Cutoff A: incentive=0)
    else:
        # Cutoff B: P/S already filled (1st half), set Q/T/R/U
        p_val = att_1st_late[att_row]
        s_val = att_1st_awol[att_row]
        q_val = att["tardy_hrs"]
        t_val = att["awol"]
        ws_att.cell(att_row, 17, q_val)
        ws_att.cell(att_row, 18, round(p_val + q_val, 2))
        ws_att.cell(att_row, 20, t_val)
        ws_att.cell(att_row, 21, s_val + t_val)

        # Compute incentive
        incentive = 0
        cur_lates = att.get("late_instances", [])
        prior_key = find_json_key(prior_data, sn_norm)
        prior_lates = []
        prior_awol = 0
        if prior_key:
            prior_att = prior_data[prior_key]
            prior_lates = prior_att.get("late_instances", [])
            prior_awol = prior_att.get("awol", 0)
        all_lates = list(cur_lates) + list(prior_lates)
        no_long = all(li.get("minutes", 0) <= LATE_THRESHOLD_MIN for li in all_lates)
        few = len(all_lates) <= LATE_COUNT_LIMIT
        no_awol = (att["awol"] == 0 and prior_awol == 0)
        if no_long and few and no_awol:
            incentive = INCENTIVE_AMOUNT
        # Override: 0 for resigned mid-period
        if sn_norm in {normalize_surname(x) for x in LWD_RESIGNATIONS.keys()}:
            incentive = 0
        ws_att.cell(att_row, 22, incentive)


# STEP 5: Update EMPLOYEE INFO (bank changes)
print(f"\n{'='*80}\nSTEP 5: Update EMPLOYEE INFO\n{'='*80}")
ws_ei = wb["EMPLOYEE INFO"]

def find_ei_row(surname, status_filter=None):
    sn_n = normalize_surname(surname)
    for r in range(2, 170):
        name = ws_ei.cell(r, 3).value
        if name and isinstance(name, str) and sn_n in normalize_surname(name):
            if status_filter:
                status = ws_ei.cell(r, 10).value
                if status and status_filter.lower() in str(status).lower():
                    return r
            else:
                return r
    return None

for sn, new_bank in BANK_UPDATES.items():
    r = find_ei_row(sn, "Permanent") or find_ei_row(sn)
    if r:
        ws_ei.cell(r, 13, new_bank)
        print(f"  {sn} (EI row {r}): bank -> {new_bank}")


# STEP 6: Update Payroll
print(f"\n{'='*80}\nSTEP 6: Update Payroll\n{'='*80}")
ws_p = wb[NEW_SHEET]

gt_row = None
for r in range(9, 55):
    for c in [2, 3, 4]:
        v = ws_p.cell(r, c).value
        if v and isinstance(v, str) and "Grand Total" in v:
            gt_row = r
            break
    if gt_row:
        break
if gt_row is None:
    gt_row = 46

# 6a. Zero resigned (full)
for sn, target_row in RESIGNED_ROW_MAP.items():
    if sn in payroll_surname_to_row:
        target_row = payroll_surname_to_row[sn]
    for c in [8, 9, 14, 16, 17, 20, 21, 22, 23]:
        ws_p.cell(target_row, c, 0)

for sn in ALREADY_ZEROED:
    if sn in payroll_surname_to_row:
        r = payroll_surname_to_row[sn]
        for c in [8, 9, 14, 16, 17, 20, 21, 22, 23]:
            ws_p.cell(r, c, 0)

# 6b. Clear period-specific cols
for r in range(9, gt_row):
    for c in [23, 26, 28, 31, 32, 36, 38]:  # W, Z, AB, AE, AF, AJ, AL
        ws_p.cell(r, c, 0)

# 6c. Adjustments (W)
for sn, amount in ADJUSTMENTS.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            ws_p.cell(pay_row, 23, amount)
            print(f"  W adj {sn} (row {pay_row}): {amount}")
            break

# 6d. Cash advances (AF)
for sn, amount in CASH_ADVANCES.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            ws_p.cell(pay_row, 32, amount)
            print(f"  AF cash {sn} (row {pay_row}): {amount}")
            break

# 6e. SSS Loans (Z)
for sn, amount in SSS_LOANS.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            ws_p.cell(pay_row, 26, amount)
            print(f"  Z SSS-loan {sn} (row {pay_row}): {amount}")
            break

# 6f. HDMF Loans (AB)
for sn, amount in HDMF_LOANS.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            ws_p.cell(pay_row, 28, amount)
            print(f"  AB HDMF-loan {sn} (row {pay_row}): {amount}")
            break

# 6g. Holiday premium
for sn, hol_info in HOLIDAY_PREMIUM.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            if "regular" in hol_info:
                ws_p.cell(pay_row, 36, hol_info["regular"])
            if "special" in hol_info:
                ws_p.cell(pay_row, 38, hol_info["special"])
            break

# 6h. Restore V formulas for hardcoded rows
for r in V_FORMULA_ROWS:
    v_formula = (
        f"=IFERROR(IF(ATTENDANCE!$D$4<>\"01-15\","
        f"(VLOOKUP(C{r},'EMPLOYEE INFO'!C:Y,21,FALSE)"
        f"+(VLOOKUP(C{r},'EMPLOYEE INFO'!C:Y,23,FALSE)*(I{r}-Q{r}))),"
        f"(VLOOKUP(C{r},'EMPLOYEE INFO'!C:Y,23,FALSE)*(I{r}-Q{r})))"
        f"+VLOOKUP(C{r},'EMPLOYEE INFO'!C:Y,20,FALSE),0)"
    )
    ws_p.cell(r, 22, v_formula)

# 6i. V overrides (mid-period resignations, special prorations)
for sn, v_value in V_OVERRIDES.items():
    sn_n = normalize_surname(sn)
    for pay_sn, pay_row in payroll_surname_to_row.items():
        if normalize_surname(pay_sn) == sn_n:
            ws_p.cell(pay_row, 22, v_value)
            print(f"  V override {sn} (row {pay_row}): {v_value}")
            break

# 6j. Habaradas days
if "habaradas" in payroll_surname_to_row:
    ws_p.cell(payroll_surname_to_row["habaradas"], 9, PERIOD_DAYS)

# 6k. Update header dates — manually verify these for your period
header_replacements = [
    # ("April 25, 2026", "May 10, 2026"),
    # ("April 01 - 15", "April 16 - 30"),
]
for r in range(1, 8):
    for c in range(1, 50):
        v = ws_p.cell(r, c).value
        if v and isinstance(v, str):
            new_v = v
            for old, new in header_replacements:
                new_v = new_v.replace(old, new)
            if new_v != v:
                ws_p.cell(r, c, new_v)


# STEP 7: Bank Details
print(f"\n{'='*80}\nSTEP 7: Bank Details\n{'='*80}")
if "Bank Details" in wb.sheetnames:
    ws_bd = wb["Bank Details"]
    if bd_oresco_row:
        ws_bd.cell(bd_oresco_row, 4, 0)
    if bd_total_row:
        ws_bd.cell(bd_total_row, 4, f"=SUM(D2:D{bd_total_row - 1})")


# STEP 8: Save & Encrypt
print(f"\n{'='*80}\nSTEP 8: Save & Encrypt\n{'='*80}")
wb.save(OUTPUT_DEC)
wb.close()
result = subprocess.run(
    ['python', '-m', 'msoffcrypto', '-p', PASSWORD, '-e', OUTPUT_DEC, OUTPUT_ENC],
    capture_output=True, text=True
)
if result.returncode == 0:
    print(f"  Encrypted: {OUTPUT_ENC}")
else:
    shutil.copy2(OUTPUT_DEC, OUTPUT_ENC)

os.makedirs(os.path.dirname(OUTPUT_FINAL), exist_ok=True)
shutil.copy2(OUTPUT_ENC, OUTPUT_FINAL)
print(f"  Final: {OUTPUT_FINAL}")
print(f"\nNow run: python verify_payroll.py --payroll {OUTPUT_DEC} \\")
print(f"  --attendance-json {ATTENDANCE_JSON} --prior-json {PRIOR_JSON} \\")
print(f"  --sheet \"{NEW_SHEET}\"")
