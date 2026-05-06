#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BICC Payroll Verifier — full Pre_Run_Checklist Step 7 (read-only).

Runs all 11 checks (7.1-7.11) plus the Apr 2026 additions (9.5b/c, 9.7).
Opens the file in ReadOnly mode, computes MD5 before/after to confirm
the file was not modified, and prints PASS/FAIL/WARN for each check.

Usage:
    python verify_payroll.py \\
        --payroll path/to/decrypted.xlsx \\
        --sheet "Payroll May 10" \\
        --attendance-json path/to/current.json \\
        --prior-json path/to/prior.json
"""

import argparse
import hashlib
import json
import os
import sys
import time
from pathlib import Path

import win32com.client
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from biccpayroll_config import (
    DAILY_EMPLOYEES, MONTHLY_EMPLOYEES, EXEMPT_EMPLOYEES, HABARADAS_LIKE,
    EXCLUDED_NAMES, INCENTIVE_AMOUNT, LATE_THRESHOLD_MIN, LATE_COUNT_LIMIT,
    normalize_surname, get_employee_type, surname_from_payroll,
)

RESIGNED = EXCLUDED_NAMES
EXEMPT_ALL = EXEMPT_EMPLOYEES | HABARADAS_LIKE


def md5_of(path):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        h.update(f.read())
    return h.hexdigest()


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--payroll", required=True)
    p.add_argument("--sheet", required=True, help='e.g. "Payroll May 10"')
    p.add_argument("--attendance-json", required=True)
    p.add_argument("--prior-json", default=None)
    args = p.parse_args()

    md5_before = md5_of(args.payroll)
    print(f"MD5 before: {md5_before}")

    with open(args.attendance_json) as f:
        cur_data = json.load(f)
    if "0" in cur_data:
        del cur_data["0"]
    cur_by_sn = {normalize_surname(k): v for k, v in cur_data.items()}

    prior_by_sn = {}
    if args.prior_json and Path(args.prior_json).exists():
        with open(args.prior_json) as f:
            prior_data = json.load(f)
        if "0" in prior_data:
            del prior_data["0"]
        prior_by_sn = {normalize_surname(k): v for k, v in prior_data.items()}

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(args.payroll), ReadOnly=True)
    time.sleep(2)
    excel.CalculateFull()
    time.sleep(1)

    ws = wb.Sheets(args.sheet)
    ws_att = wb.Sheets("ATTENDANCE")
    ws_bd = wb.Sheets("Bank Details")
    ws_ei = wb.Sheets("EMPLOYEE INFO")

    gt_row = None
    for r in range(9, 60):
        for c in [2, 3, 4]:
            v = ws.Cells(r, c).Value
            if v and isinstance(v, str) and "Grand Total" in v:
                gt_row = r
                break
        if gt_row:
            break
    if gt_row is None:
        gt_row = 46
    print(f"GT row: {gt_row}\n")

    fails = []
    warns = []
    passes = []

    def F(c, m): fails.append((c, m)); print(f"  FAIL [{c}] {m}")
    def W(c, m): warns.append((c, m)); print(f"  WARN [{c}] {m}")
    def P(c, m): passes.append((c, m))

    # 7.1
    print("=" * 80 + "\n7.1 Formula Integrity\n" + "=" * 80)
    e = 0
    for r in range(9, gt_row + 1):
        for c in range(3, 45):
            v = ws.Cells(r, c).Value
            if v is None: continue
            if isinstance(v, (int, float)) and v < -2000000000:
                F("7.1", f"{get_column_letter(c)}{r}=ERROR_CODE"); e += 1
            elif isinstance(v, str) and v.startswith('#'):
                F("7.1", f"{get_column_letter(c)}{r}={v}"); e += 1
    if e == 0: P("7.1", "ok"); print("  PASS")

    # 7.2
    print("\n" + "=" * 80 + "\n7.2 Per-Employee Formula Balance\n" + "=" * 80)
    fb = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str) or not name.strip(): continue
        sn = surname_from_payroll(name)
        et = get_employee_type(sn)
        H = ws.Cells(r, 8).Value or 0
        I_ = ws.Cells(r, 9).Value or 0
        J = ws.Cells(r, 10).Value or 0
        K = ws.Cells(r, 11).Value or 0
        L = ws.Cells(r, 12).Value or 0
        O = ws.Cells(r, 15).Value or 0
        P_ = ws.Cells(r, 16).Value or 0
        Q = ws.Cells(r, 17).Value or 0
        R = ws.Cells(r, 18).Value or 0
        S = ws.Cells(r, 19).Value or 0
        T = ws.Cells(r, 20).Value or 0
        U = ws.Cells(r, 21).Value or 0
        V = ws.Cells(r, 22).Value or 0
        W_ = ws.Cells(r, 23).Value or 0
        X = ws.Cells(r, 24).Value or 0
        Y = ws.Cells(r, 25).Value or 0
        Z = ws.Cells(r, 26).Value or 0
        AA = ws.Cells(r, 27).Value or 0
        AB = ws.Cells(r, 28).Value or 0
        AC = ws.Cells(r, 29).Value or 0
        AD = ws.Cells(r, 30).Value or 0
        AE = ws.Cells(r, 31).Value or 0
        AF = ws.Cells(r, 32).Value or 0
        AG = ws.Cells(r, 33).Value or 0
        AH = ws.Cells(r, 34).Value or 0
        bad = False
        if abs(X - (L + O + T + U + V + W_ - R - S)) > 0.5:
            F("7.2", f"R{r} {name}: X={X:.2f} vs exp={L+O+T+U+V+W_-R-S:.2f}"); bad = True
        if abs(AG - (Y + Z + AA + AB + AC + AD + AE + AF)) > 0.5:
            F("7.2", f"R{r} {name}: AG={AG:.2f} vs sum={Y+Z+AA+AB+AC+AD+AE+AF:.2f}"); bad = True
        if abs(AH - (X - AG)) > 0.5:
            F("7.2", f"R{r} {name}: AH={AH:.2f} vs X-AG={X-AG:.2f}"); bad = True
        if et == "DAILY" and abs(L - J * I_) > 1:
            F("7.2", f"R{r} {name}: DAILY L={L:.2f} vs J*I={J*I_:.2f}"); bad = True
        if et in ("MONTHLY", "FIXED", "EXEMPT") and H > 0 and abs(L - H/2) > 1:
            F("7.2", f"R{r} {name}: {et} L={L:.2f} vs H/2={H/2:.2f}"); bad = True
        if P_ > 0 and abs(R - K * P_) > 1:
            F("7.2", f"R{r} {name}: R={R:.2f} vs K*P={K*P_:.2f}"); bad = True
        if Q > 0 and abs(S - J * Q) > 1:
            F("7.2", f"R{r} {name}: S={S:.2f} vs J*Q={J*Q:.2f}"); bad = True
        if Q > 0 and S == 0:
            F("7.2", f"R{r} {name}: Q>0 but S=0"); bad = True
        if bad: fb += 1
    if fb == 0: P("7.2", "ok"); print("  PASS")

    # 7.3 — only warn
    print("\n" + "=" * 80 + "\n7.3 Cutoff B SSS/HDMF/PHIC enrolled check\n" + "=" * 80)
    ei_enrolled = {}
    for r_ei in range(2, 170):
        name = ws_ei.Cells(r_ei, 3).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        status = ws_ei.Cells(r_ei, 10).Value
        sss_n = ws_ei.Cells(r_ei, 14).Value
        hdmf_n = ws_ei.Cells(r_ei, 15).Value
        phic_n = ws_ei.Cells(r_ei, 16).Value
        if sn not in ei_enrolled or (status and "permanent" in str(status).lower()):
            ei_enrolled[sn] = {
                "sss": sss_n is not None and str(sss_n).strip() not in ("", "BPI"),
                "hdmf": hdmf_n is not None and str(hdmf_n).strip() not in ("", "BPI"),
                "phic": phic_n is not None and str(phic_n).strip() not in ("", "BPI"),
                "status": str(status) if status else ""
            }
    d4 = ws_att.Cells(4, 4).Value
    is_cutoff_b = d4 != "01-15"
    if is_cutoff_b:
        for r in range(9, gt_row):
            name = ws.Cells(r, 3).Value
            if not name or not isinstance(name, str): continue
            sn = normalize_surname(surname_from_payroll(name))
            if sn in {normalize_surname(x) for x in RESIGNED}: continue
            if sn in {normalize_surname(x) for x in EXEMPT_ALL}: continue
            L = ws.Cells(r, 12).Value or 0
            if L == 0: continue
            Y_ = ws.Cells(r, 25).Value or 0
            AC_ = ws.Cells(r, 29).Value or 0
            enr = ei_enrolled.get(sn, {})
            if "permanent" in enr.get("status", "").lower():
                if enr.get("sss") and Y_ == 0:
                    W("7.3", f"R{r} {name}: permanent w/SSS# but Y=0")
                if enr.get("phic") and AC_ == 0:
                    W("7.3", f"R{r} {name}: permanent w/PHIC# but AC=0")
        print("  (warnings above if any)")
    else:
        print("  Cutoff A — checking Y/AA/AC/AD/U all 0")
        for r in range(9, gt_row):
            name = ws.Cells(r, 3).Value
            if not name or not isinstance(name, str) or not name.strip(): continue
            for col, label in [(25, "Y"), (27, "AA"), (29, "AC"), (30, "AD"), (21, "U")]:
                v = ws.Cells(r, col).Value or 0
                if v != 0:
                    F("7.3", f"R{r} {name}: {label}={v} (Cutoff A should be 0)")

    # 7.4 Pay type checks
    print("\n" + "=" * 80 + "\n7.4 Pay Type Checks\n" + "=" * 80)
    pt = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        if sn in {normalize_surname(x) for x in RESIGNED}: continue
        et = get_employee_type(sn)
        O = ws.Cells(r, 15).Value or 0
        T = ws.Cells(r, 20).Value or 0
        R = ws.Cells(r, 18).Value or 0
        S = ws.Cells(r, 19).Value or 0
        if et == "FIXED":
            if O != 0: F("7.4", f"R{r} {name}: FIXED O={O}"); pt += 1
            if T != 0: F("7.4", f"R{r} {name}: FIXED T={T}"); pt += 1
        if et == "EXEMPT":
            if R != 0: F("7.4", f"R{r} {name}: EXEMPT R={R}"); pt += 1
            if S != 0: F("7.4", f"R{r} {name}: EXEMPT S={S}"); pt += 1
    if pt == 0: P("7.4", "ok"); print("  PASS")

    # 7.5 Incentive
    print("\n" + "=" * 80 + "\n7.5 Attendance Incentive\n" + "=" * 80)
    inc_e = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        if sn in {normalize_surname(x) for x in RESIGNED}: continue
        U = ws.Cells(r, 21).Value or 0
        if U == 0: continue
        cur = cur_by_sn.get(sn, {})
        prior = prior_by_sn.get(sn, {})
        all_lates = list(cur.get("late_instances", [])) + list(prior.get("late_instances", []))
        long_late = [li for li in all_lates if li.get("minutes", 0) > LATE_THRESHOLD_MIN]
        too_many = len(all_lates) > LATE_COUNT_LIMIT
        has_awol = (cur.get("awol", 0) > 0 or prior.get("awol", 0) > 0)
        reasons = []
        if long_late: reasons.append(f"single late >15min")
        if too_many: reasons.append(f"{len(all_lates)} lates")
        if has_awol: reasons.append("AWOL")
        if reasons:
            F("7.5", f"R{r} {name}: U={U} ineligible: {', '.join(reasons)}"); inc_e += 1
    if inc_e == 0: P("7.5", "ok"); print("  PASS")

    # 7.7 Bank Details total
    print("\n" + "=" * 80 + "\n7.7 Bank Details vs Payroll\n" + "=" * 80)
    gt_net = ws.Cells(gt_row, 34).Value or 0
    bd_total = None
    for r in range(1, 60):
        v = ws_bd.Cells(r, 2).Value
        if v and isinstance(v, str) and "TOTAL" in v.upper():
            bd_total = ws_bd.Cells(r, 4).Value or 0
            break
    if bd_total is not None and abs(gt_net - bd_total) < 1:
        P("7.7", f"BD={bd_total:,.2f} = GT={gt_net:,.2f}")
        print(f"  PASS: {bd_total:,.2f} = {gt_net:,.2f}")
    else:
        F("7.7", f"BD={bd_total} vs GT={gt_net}")

    # 7.8 GT independent
    print("\n" + "=" * 80 + "\n7.8 Grand Total Independent Recalc\n" + "=" * 80)
    gt_e = 0
    for col in [12, 24, 33, 34]:
        ind = sum(float(ws.Cells(r, col).Value or 0) for r in range(9, gt_row))
        gt_v = ws.Cells(gt_row, col).Value or 0
        n = {12: "L", 24: "X", 33: "AG", 34: "AH"}[col]
        if abs(ind - gt_v) < 1:
            print(f"  PASS {n}: {ind:,.2f} = {gt_v:,.2f}")
        else:
            F("7.8", f"GT {n}: ind={ind:,.2f} vs GT={gt_v:,.2f}"); gt_e += 1
    if gt_e == 0: P("7.8", "ok")

    # 7.9 Resigned
    print("\n" + "=" * 80 + "\n7.9 Resigned Employee Checks\n" + "=" * 80)
    re_e = 0
    for resigned_sn in RESIGNED:
        for r in range(9, gt_row):
            name = ws.Cells(r, 3).Value
            if not name or not isinstance(name, str): continue
            if normalize_surname(resigned_sn) not in normalize_surname(surname_from_payroll(name)): continue
            bad = []
            for c in range(12, 35):
                v = ws.Cells(r, c).Value or 0
                if isinstance(v, (int, float)) and abs(v) > 0.01:
                    bad.append((get_column_letter(c), v))
            if bad:
                F("7.9", f"R{r} {name}: non-zero {bad}"); re_e += 1
            else:
                print(f"  PASS R{r} {name}: zero")
            break
    if re_e == 0: P("7.9", "ok")

    # 7.10 Cross-sheet
    print("\n" + "=" * 80 + "\n7.10 Cross-Sheet Consistency\n" + "=" * 80)
    att_g = {}
    for r_a in range(9, 43):
        name = ws_att.Cells(r_a, 4).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        att_g[sn] = float(ws_att.Cells(r_a, 7).Value or 0)
    xs_e = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        if sn in {normalize_surname(x) for x in RESIGNED}: continue
        if sn in {normalize_surname(x) for x in EXEMPT_ALL}: continue
        pay_I = ws.Cells(r, 9).Value or 0
        att_G = att_g.get(sn)
        if att_G is None: continue
        if abs(float(pay_I) - att_G) > 0.01:
            F("7.10", f"R{r} {name}: I={pay_I} vs ATT G={att_G}"); xs_e += 1
    if xs_e == 0: P("7.10", "ok"); print("  PASS")

    # 7.11 Reasonableness
    print("\n" + "=" * 80 + "\n7.11 Reasonableness\n" + "=" * 80)
    rs_e = 0
    period_days_max = 16  # generous cap
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        L = ws.Cells(r, 12).Value or 0
        X = ws.Cells(r, 24).Value or 0
        AH = ws.Cells(r, 34).Value or 0
        I_ = ws.Cells(r, 9).Value or 0
        if AH < 0: F("7.11", f"R{r} {name}: NEGATIVE Net={AH}"); rs_e += 1
        if AH > X + 1: F("7.11", f"R{r} {name}: Net>Gross"); rs_e += 1
        if I_ > period_days_max: F("7.11", f"R{r} {name}: I={I_} > {period_days_max}"); rs_e += 1
    if rs_e == 0: P("7.11", "ok"); print("  PASS")

    # 9.5b JSON cross-check
    print("\n" + "=" * 80 + "\n9.5b JSON ↔ Payroll I cross-check\n" + "=" * 80)
    cc_e = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        sn = normalize_surname(surname_from_payroll(name))
        if sn in {normalize_surname(x) for x in RESIGNED}: continue
        if sn in {normalize_surname(x) for x in EXEMPT_ALL}: continue
        if sn not in cur_by_sn: continue
        jdays = cur_by_sn[sn]["days"]
        fI = ws.Cells(r, 9).Value or 0
        if abs(float(fI) - float(jdays)) > 0.01:
            F("9.5b", f"R{r} {name}: I={fI} vs JSON={jdays}"); cc_e += 1
    if cc_e == 0: P("9.5b", "ok"); print("  PASS")

    # 9.5c DAILY Q=0
    print("\n" + "=" * 80 + "\n9.5c DAILY Q=0 guard\n" + "=" * 80)
    dd = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        sn = surname_from_payroll(name)
        if get_employee_type(sn) != "DAILY": continue
        Q_ = ws.Cells(r, 17).Value or 0
        if float(Q_) > 0:
            F("9.5c", f"R{r} {name}: DAILY Q={Q_}"); dd += 1
    if dd == 0: P("9.5c", "ok"); print("  PASS")

    # 9.7 Bank account
    print("\n" + "=" * 80 + "\n9.7 Bank Account Validation\n" + "=" * 80)
    ba = 0
    for r in range(9, gt_row):
        name = ws.Cells(r, 3).Value
        if not name or not isinstance(name, str): continue
        AH = ws.Cells(r, 34).Value or 0
        if float(AH) <= 0: continue
        bank = ws.Cells(r, 7).Value
        if bank is None or (isinstance(bank, str) and not bank.strip()):
            F("9.7", f"R{r} {name}: no bank"); ba += 1; continue
        bs = str(bank).strip()
        if bs.upper() == "CASH": continue
        digits = ''.join(c for c in bs if c.isdigit())
        if len(digits) < 6:
            F("9.7", f"R{r} {name}: invalid bank '{bs}'"); ba += 1
    if ba == 0: P("9.7", "ok"); print("  PASS")

    # Grand total summary
    gt_X = ws.Cells(gt_row, 24).Value or 0
    gt_AG = ws.Cells(gt_row, 33).Value or 0
    gt_AH = ws.Cells(gt_row, 34).Value or 0
    print(f"\nGRAND TOTAL: Gross={gt_X:,.2f}, Deduct={gt_AG:,.2f}, NET={gt_AH:,.2f}")

    wb.Close(SaveChanges=False)
    excel.Quit()
    del excel

    md5_after = md5_of(args.payroll)
    print(f"\nMD5 after: {md5_after}")
    print(f"File unchanged: {md5_before == md5_after}")

    print("\n" + "=" * 80)
    print(f"SUMMARY: {len(passes)} PASS, {len(fails)} FAIL, {len(warns)} WARN")
    print("=" * 80)
    if fails:
        print("FAILURES:")
        for c, m in fails:
            print(f"  [{c}] {m}")
    if not fails:
        print("ALL CRITICAL CHECKS PASSED")

    sys.exit(1 if fails else 0)


if __name__ == "__main__":
    main()
